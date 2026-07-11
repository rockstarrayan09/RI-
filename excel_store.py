import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from config import DATA_DIR, EXCEL_PATH

HEADERS = [
    "Date & Time",
    "Service Name",
    "Amount",
    "Customer Name",
    "Mobile",
    "Aadhaar",
    "Status",
    "Reference Number",
    "Government Amount",
    "Profit",
]

DEFAULT_STATUS = "Process"
DEFAULT_REFERENCE_NUMBER = "NA"


def validate_submission(name, mobile, aadhaar):
    errors = []
    name = (name or "").strip()
    mobile = (mobile or "").strip()
    aadhaar = (aadhaar or "").strip() or "123456789000"

    if len(name) < 2:
        errors.append("Customer name must be at least 2 characters.")
    if not re.fullmatch(r"\d{10}", mobile):
        errors.append("Mobile number must be exactly 10 digits.")
    if not re.fullmatch(r"\d{12}", aadhaar):
        errors.append("Aadhaar number must be exactly 12 digits.")

    return errors, name, mobile, aadhaar


def _ensure_workbook():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if not EXCEL_PATH.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Records"
        ws.append(HEADERS)
        for col in range(1, len(HEADERS) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)
        wb.save(EXCEL_PATH)
    else:
        _migrate_workbook_if_needed()


def ensure_workbook_ready():
    """Public entry point to guarantee the Excel file is fully migrated
    (headers + backfilled columns) before it's read or downloaded."""
    _ensure_workbook()


def _migrate_workbook_if_needed():
    """Backfill older records.xlsx files that were created before the
    Status / Reference Number / Government Amount / Profit columns existed.
    Without this, those columns stay blank forever on old rows."""
    if not EXCEL_PATH.exists():
        return

    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    changed = False

    # Pad the header row if it's shorter than the current HEADERS list,
    # and make sure every header cell has the right text and is bold.
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        if cell.value != HEADERS[col - 1]:
            cell.value = HEADERS[col - 1]
            changed = True
        if not (cell.font and cell.font.bold):
            cell.font = Font(bold=True)
            changed = True

    amount_col = HEADERS.index("Amount") + 1
    status_col = HEADERS.index("Status") + 1
    reference_col = HEADERS.index("Reference Number") + 1
    government_col = HEADERS.index("Government Amount") + 1
    profit_col = HEADERS.index("Profit") + 1

    for row in range(2, ws.max_row + 1):
        service_name = ws.cell(row=row, column=2).value
        if not service_name:
            continue

        if not ws.cell(row=row, column=status_col).value:
            ws.cell(row=row, column=status_col, value=DEFAULT_STATUS)
            changed = True

        if not ws.cell(row=row, column=reference_col).value:
            ws.cell(row=row, column=reference_col, value=DEFAULT_REFERENCE_NUMBER)
            changed = True

        government_amount = ws.cell(row=row, column=government_col).value
        if government_amount in (None, ""):
            government_amount = 0
            ws.cell(row=row, column=government_col, value=government_amount)
            changed = True

        profit = ws.cell(row=row, column=profit_col).value
        if profit in (None, ""):
            amount = ws.cell(row=row, column=amount_col).value or 0
            ws.cell(row=row, column=profit_col, value=amount - government_amount)
            changed = True

    if changed:
        wb.save(EXCEL_PATH)


def get_all_records():
    _ensure_workbook()
    if not EXCEL_PATH.exists():
        return []

    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    records = []
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[1]:  # Service Name column empty -> treat as blank row
            continue
        record = dict(zip(HEADERS, row))
        record["row_number"] = row_number  # needed to edit/delete this exact row later
        records.append(record)

    records.reverse()  # newest submissions first
    return records


def get_record(row_number):
    for record in get_all_records():
        if record.get("row_number") == row_number:
            return record
    return None


def update_record(
    row_number,
    service_name,
    amount,
    customer_name,
    mobile,
    aadhaar,
    status=None,
    reference_number=None,
    government_amount=0,
):
    """Overwrite an existing row in place (identified by row_number, which
    comes from get_all_records()). The original Date & Time is preserved."""
    if not EXCEL_PATH.exists():
        return False

    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    if row_number < 2 or row_number > ws.max_row:
        return False
    if not ws.cell(row=row_number, column=2).value:
        return False  # row is blank / already deleted

    amount = amount or 0
    government_amount = government_amount or 0
    profit = amount - government_amount

    values = [
        ws.cell(row=row_number, column=1).value,  # keep original Date & Time
        service_name,
        amount,
        customer_name,
        mobile,
        aadhaar,
        (status or "").strip() or DEFAULT_STATUS,
        (reference_number or "").strip() or DEFAULT_REFERENCE_NUMBER,
        government_amount,
        profit,
    ]
    for col, value in enumerate(values, start=1):
        ws.cell(row=row_number, column=col, value=value)

    wb.save(EXCEL_PATH)
    return True


def delete_record(row_number):
    """Blanks out the row so it's excluded from get_all_records() and its
    slot is naturally reused by append_record next time (same pattern the
    app already uses for 'first empty row')."""
    if not EXCEL_PATH.exists():
        return False

    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    if row_number < 2 or row_number > ws.max_row:
        return False

    for col in range(1, len(HEADERS) + 1):
        ws.cell(row=row_number, column=col).value = None

    wb.save(EXCEL_PATH)
    return True


def _parse_record_datetime(value):
    """Records normally store Date & Time as text like '2026-07-10 08:02:30 PM'.
    But if the file is ever opened and re-saved in Excel/Google Sheets, that
    text can get silently converted into a real datetime value. Handle both,
    so records never silently vanish from the monthly report."""
    if isinstance(value, datetime):
        return value
    if not value:
        return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %I:%M:%S %p", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def get_monthly_report():
    """Groups all records by month, returning total sales/profit/government
    amount per month, a per-service breakdown (sorted so the top-selling
    service is first), and an all-time grand total across every record."""
    records = get_all_records()
    monthly = defaultdict(
        lambda: {
            "sales": 0.0,
            "profit": 0.0,
            "government_amount": 0.0,
            "count": 0,
            "services": defaultdict(
                lambda: {"sales": 0.0, "profit": 0.0, "government_amount": 0.0, "count": 0}
            ),
        }
    )

    overall = {"sales": 0.0, "profit": 0.0, "government_amount": 0.0, "count": 0}

    for record in records:
        dt = _parse_record_datetime(record.get("Date & Time"))
        if dt is None:
            continue

        amount = record.get("Amount") or 0
        government_amount = record.get("Government Amount") or 0
        profit = record.get("Profit")
        if profit in (None, ""):
            profit = amount - government_amount
        service_name = record.get("Service Name") or "Unknown"

        overall["sales"] += amount
        overall["profit"] += profit
        overall["government_amount"] += government_amount
        overall["count"] += 1

        month_key = dt.strftime("%Y-%m")
        month_data = monthly[month_key]
        month_data["sales"] += amount
        month_data["profit"] += profit
        month_data["government_amount"] += government_amount
        month_data["count"] += 1

        service_data = month_data["services"][service_name]
        service_data["sales"] += amount
        service_data["profit"] += profit
        service_data["government_amount"] += government_amount
        service_data["count"] += 1

    months = []
    for month_key in sorted(monthly.keys(), reverse=True):
        data = monthly[month_key]
        services_list = [
            {"name": name, **stats} for name, stats in data["services"].items()
        ]
        services_list.sort(key=lambda s: s["sales"], reverse=True)

        months.append(
            {
                "month": month_key,
                "month_label": datetime.strptime(month_key, "%Y-%m").strftime("%B %Y"),
                "sales": data["sales"],
                "profit": data["profit"],
                "government_amount": data["government_amount"],
                "count": data["count"],
                "top_service": services_list[0]["name"] if services_list else None,
                "services": services_list,
            }
        )

    return {"overall": overall, "months": months}


def append_record(
    service_name,
    amount,
    customer_name,
    mobile,
    aadhaar,
    government_amount=0,
):
    _ensure_workbook()
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    government_amount = government_amount or 0
    profit = amount - government_amount

    data = [
        datetime.now().strftime("%Y-%m-%d %I:%M:%S %p"),
        service_name,
        amount,
        customer_name,
        mobile,
        aadhaar,
        DEFAULT_STATUS,
        DEFAULT_REFERENCE_NUMBER,
        government_amount,
        profit,
    ]

    # Find first empty row
    target_row = None

    for row in range(2, ws.max_row + 1):
        if not ws.cell(row=row, column=2).value:  # Service Name column
            target_row = row
            break

    if target_row is None:
        target_row = ws.max_row + 1

    for col, value in enumerate(data, start=1):
        ws.cell(row=target_row, column=col, value=value)

    wb.save(EXCEL_PATH)